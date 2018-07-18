# coding=utf-8
import os
import shutil
import zipfile
import openpyxl
from openpyxl.drawing.image import Image


class ExcelImgRead(object):

    def copy_file_to_zip(self, file_path, old_name, new_type='.zip'):
        """
        复制指定目录下的文件并修改类型为'.zip'
        :param file_path:
        :param old_name:
        :param new_type:
        :return:
        """
        old_path = os.path.join(file_path, old_name)
        if not os.path.exists(old_path):
            print ('No such File! :%s' % old_path)
            return False
        new_name = str(old_name.split('.')[0]) + new_type
        new_path = os.path.join(file_path, new_name)
        if os.path.exists(new_path):
            os.remove(new_path)
        # os.rename(old_path, new_path)
        shutil.copyfile(old_path, new_path)

    def unzip_file(self, file_path):
        """
        解压缩指定目录下的Zip文件
        :param file_path:
        :return:
        """
        file_list = os.listdir(file_path)
        for file_name in file_list:
            if file_name.split('.')[-1] == 'zip':
                try:
                    file_abs_path = os.path.join(file_path, file_name)
                    target_dir = file_name.split('.')[0]
                    target_abs_dir = os.path.join(file_path, target_dir)
                    if os.path.exists(target_abs_dir):
                        shutil.rmtree(target_abs_dir)
                    zipf = zipfile.ZipFile(file_abs_path, 'r')
                    zipf.extractall(target_abs_dir)  # 解压到指定文件目录
                    zipf.close()
                    os.remove(file_abs_path)
                except:
                    print 'Unzip Zipfile Failed '
                    shutil.rmtree(target_abs_dir)

    def get_excel_pic(self, file_path, file_name):
        """
        解压缩的excel目录下获取图片
        :param file_path:
        :param file_name:
        :return list():
        """
        pic_dir = 'xl\media'
        pic_path = os.path.join(file_path, file_name.split('.')[0], pic_dir)
        if not os.path.exists(pic_path):
            print ('No such directory!:%s' % pic_path)
            return ''
        file_list = os.listdir(pic_path)
        paths = []
        for file_name in file_list:
            if file_name.split('.')[1] == 'png':
                path = os.path.join(pic_path, file_name)
                paths.append(path)
        return paths


def excel_pic_read(file_path, file_name):
    """
    读取excel中的图片
    :param file_path:
    :param file_name:
    :return:图片的path
    """
    eir = ExcelImgRead()
    eir.copy_file_to_zip(file_path, file_name)
    eir.unzip_file(file_path)
    return eir.get_excel_pic(file_path, file_name)


if __name__ == '__main__':
    eir = ExcelImgRead()
    file_path = r"C:\Users\sang\PycharmProjects\PythonByte\handle_excel\files"
    file_name = 'test_img.xlsx'
    paths = excel_pic_read(file_path, file_name)

    # workbook = openpyxl.load_workbook(
    #     os.path.join(file_path, file_name))
    # worksheet = workbook['BUR']
    # # workbook.create_sheet(title=, index=)
    #
    # row = worksheet.max_row
    # for path in paths:
    #     img = Image(path)
    #     pos = 'A' + str(row)  # worksheet.max_row取得的行数没有包含图片占的行，所以add_image()之后max_row没有改变，图片会重叠到一起
    #     worksheet.add_image(img, pos)
    #     row += img.height / 20 + 2
    # workbook.save(os.path.join(file_path, 'new_img.xlsx'))
