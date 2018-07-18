# coding=utf-8
import openpyxl
from openpyxl.worksheet.copier import WorksheetCopy

workbook = openpyxl.load_workbook('test_img.xlsx')  # 读取xlsx
template_worksheet = workbook['BUR']

new_worksheet = workbook.create_sheet('New_Sheet_Name')

instance = WorksheetCopy(template_worksheet, new_worksheet)
WorksheetCopy.copy_worksheet(instance)
# wb.create_sheet('test', index=0)
workbook.save('new_wb.xlsx')
